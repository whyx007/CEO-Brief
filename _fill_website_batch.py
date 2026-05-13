# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from urllib.parse import quote, urlparse
from urllib.request import Request, urlopen
import re, time, sys
from pathlib import Path

infile = Path(r'C:\Users\Administrator\.openclaw\workspace\finance-postinvest-ai-platform\ceo-brief\company-info\company-info_with_url.xlsx')
wb = load_workbook(infile)
ws = wb[wb.sheetnames[0]]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
url_col = next((i for i,h in enumerate(headers, start=1) if str(h).strip() == '官网URL'), None)
if url_col is None:
    raise RuntimeError(f'官网URL not found: {headers}')

ua = {'User-Agent': 'Mozilla/5.0'}
pat = re.compile(r'<a[^>]+class="result__a"[^>]+href="([^"]+)"[^>]*>(.*?)</a>', re.S)
BAD = ['baidu.com','bing.com','so.com','sogou.com','zhihu.com','bilibili.com','weixin.qq.com','xiaohongshu.com','weibo.com','douyin.com','google.com']

def pick_official(name):
    q = quote(f'{name} 官网')
    url = f'https://duckduckgo.com/html/?q={q}'
    req = Request(url, headers=ua)
    html = urlopen(req, timeout=15).read().decode('utf-8', 'ignore')
    for href, _title in pat.findall(html):
        if href.startswith('//'):
            href = 'https:' + href
        if href.startswith('/l/?'):
            continue
        host = urlparse(href).netloc.lower()
        if not host:
            continue
        if any(b in host for b in BAD):
            continue
        return href
    return ''

start = 2
end = min(ws.max_row, 21)
count = 0
for row_idx in range(start, end + 1):
    name = str(ws.cell(row_idx, 1).value or '').strip()
    if not name:
        continue
    try:
        url = pick_official(name)
    except Exception as e:
        url = ''
        print('ERR', row_idx, name, repr(e), flush=True)
    ws.cell(row_idx, url_col).value = url
    count += 1
    print(f'{row_idx}: {name} => {url}', flush=True)
    time.sleep(0.4)

wb.save(infile)
print(f'BATCH_DONE {count}', flush=True)
