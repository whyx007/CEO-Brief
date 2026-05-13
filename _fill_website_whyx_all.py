# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from urllib.parse import quote, urlparse, parse_qs, unquote
from urllib.request import Request, urlopen
from html import unescape
import re, time
from pathlib import Path

infile = Path(r'C:\Users\Administrator\.openclaw\workspace\finance-postinvest-ai-platform\ceo-brief\company-info\company-info_with_url.xlsx')
wb = load_workbook(infile)
ws = wb[wb.sheetnames[0]]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
url_col = next((i for i, h in enumerate(headers, start=1) if str(h).strip() == '官网URL'), None)
if url_col is None:
    raise RuntimeError('官网URL not found')

ua = {'User-Agent': 'Mozilla/5.0'}
BAD = [
    'baidu.com','bing.com','so.com','sogou.com','zhihu.com','bilibili.com','weixin.qq.com',
    'xiaohongshu.com','weibo.com','douyin.com','google.com','qcc.com','tianyancha.com',
    'aiqicha.baidu.com','zhaopin.com','liepin.com','58.com','kanzhun.com','jobui.com',
    'made-in-china.com','alibaba.com','1688.com','huicong.com','cnelc.com','baike.baidu.com'
]

def clean_href(href: str) -> str:
    href = unescape(href).replace('&amp;', '&')
    if href.startswith('//'):
        href = 'https:' + href
    if 'duckduckgo.com/l/?' in href:
        q = parse_qs(urlparse(href).query)
        if 'uddg' in q:
            href = unquote(q['uddg'][0])
    return href


def score_url(url: str, title: str, name: str) -> int:
    host = urlparse(url).netloc.lower()
    if not host:
        return -999
    if any(b in host for b in BAD):
        return -999
    score = 0
    if name.replace('（人工核对）','').replace('（','').replace('）','')[:6] and name[:6] in title:
        score += 20
    if host.endswith('.cn') or host.endswith('.com') or host.endswith('.vip'):
        score += 5
    if any(k in title for k in ['官网','官网首页','官方网站','首页']):
        score += 10
    if any(k in url for k in ['about','home','index','p-about','wzsy']):
        score += 3
    return score


def pick_official(name: str) -> str:
    q = quote(f'{name} 官网')
    search_url = f'https://search.whyx.site:8444/search?q={q}'
    html = urlopen(Request(search_url, headers=ua), timeout=25).read().decode('utf-8', 'ignore')
    candidates = []
    # result articles with h3 links are best
    for m in re.finditer(r'<article[^>]*class="[^"]*result[^"]*"[^>]*>.*?<h3><a href="([^"]+)"[^>]*>(.*?)</a>', html, re.S):
        href = clean_href(m.group(1))
        title = re.sub(r'<.*?>', '', unescape(m.group(2))).strip()
        candidates.append((score_url(href, title, name), href, title))
    # fallback: any url_header urls
    for m in re.finditer(r'<a href="([^"]+)" class="url_header"', html):
        href = clean_href(m.group(1))
        candidates.append((score_url(href, href, name) - 1, href, href))
    candidates.sort(key=lambda x: (-x[0], len(urlparse(x[1]).netloc), x[1]))
    for sc, href, title in candidates:
        if sc <= 0:
            continue
        return href
    return ''

start_row = 2
for row in range(start_row, ws.max_row + 1):
    name = str(ws.cell(row, 1).value or '').strip()
    if not name:
        continue
    current = str(ws.cell(row, url_col).value or '').strip()
    if current:
        continue
    try:
        url = pick_official(name)
    except Exception as e:
        url = ''
        print(f'ERR {row} {name} {e!r}', flush=True)
    ws.cell(row, url_col).value = url
    wb.save(infile)
    print(f'{row}: {name} => {url}', flush=True)
    time.sleep(0.35)

print('DONE', flush=True)
