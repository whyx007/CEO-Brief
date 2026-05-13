# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from urllib.parse import urlparse, parse_qs, unquote
from html import unescape
from pathlib import Path
import re

infile = Path(r'C:\Users\Administrator\.openclaw\workspace\finance-postinvest-ai-platform\ceo-brief\company-info\company-info_with_url.xlsx')
wb = load_workbook(infile)
ws = wb[wb.sheetnames[0]]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
url_col = next((i for i,h in enumerate(headers, start=1) if str(h).strip() == '官网URL'), None)
if url_col is None:
    raise RuntimeError('官网URL not found')

BAD = [
    'duckduckgo.com', 'qcc.com', 'tianyancha.com', 'aiqicha.baidu.com', 'baike.baidu.com',
    'zhaopin.com', 'liepin.com', 'jobui.com', '58.com', 'kanzhun.com', 'yingjiesheng.com',
    'iyiou.com', 'data.iyiou.com', 'm.baike.com', 'm.cphi.cn', 'smejs.cn', 'waiqicha.com',
    '11467.com', 'gsiecq.com', 'promotac.com', 'career.nankai.edu.cn', 'yesky.com',
    'qixin.com', 'cphi.cn', 'chain.pdf', 'qichacha.com'
]

# 如果是DDG跳转链，解包到uddg目标

def unwrap(url: str) -> str:
    url = unescape((url or '').strip())
    if not url:
        return ''
    if url.startswith('https://duckduckgo.com/l/?') or url.startswith('http://duckduckgo.com/l/?'):
        q = parse_qs(urlparse(url).query)
        if 'uddg' in q and q['uddg']:
            url = unquote(q['uddg'][0])
    if url.startswith('//'):
        url = 'https:' + url
    return url


def is_bad(url: str) -> bool:
    host = urlparse(url).netloc.lower()
    return any(b in host for b in BAD)

changed = 0
cleared = 0
for r in range(2, ws.max_row + 1):
    raw = str(ws.cell(r, url_col).value or '').strip()
    if not raw:
        continue
    url = unwrap(raw)
    if not url or is_bad(url):
        ws.cell(r, url_col).value = ''
        cleared += 1
        continue
    # 统一去掉尾部多余空白
    if url != raw:
        ws.cell(r, url_col).value = url
        changed += 1

wb.save(infile)
print({'changed': changed, 'cleared': cleared, 'rows': ws.max_row - 1})
