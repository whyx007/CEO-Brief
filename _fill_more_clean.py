# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from urllib.parse import quote, urlparse, unquote, parse_qs
from urllib.request import Request, urlopen
from html import unescape
from pathlib import Path
import re, time

infile = Path(r'C:\Users\Administrator\.openclaw\workspace\finance-postinvest-ai-platform\ceo-brief\company-info\company-info_with_url.xlsx')
wb = load_workbook(infile)
ws = wb[wb.sheetnames[0]]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
url_col = next((i for i, h in enumerate(headers, start=1) if str(h).strip() == '官网URL'), None)
if url_col is None:
    raise RuntimeError('官网URL not found')

UA = {'User-Agent': 'Mozilla/5.0'}
BAD = [
    'duckduckgo.com','qcc.com','tianyancha.com','aiqicha.baidu.com','baike.baidu.com',
    'zhaopin.com','liepin.com','jobui.com','58.com','kanzhun.com','yingjiesheng.com',
    'iyiou.com','data.iyiou.com','m.baike.com','m.cphi.cn','smejs.cn','waiqicha.com',
    '11467.com','gsiecq.com','promotac.com','career.nankai.edu.cn','yesky.com','qixin.com',
    'cphi.cn','c.gongkong.com','c-semt.com','csoe.org.cn','gys.cn','baidu.com','liepin.com'
]

article_pat = re.compile(r'<article[^>]*class="[^"]*result[^"]*"[^>]*>(.*?)</article>', re.S)
url_pat = re.compile(r'<h3><a href="([^"]+)"[^>]*>(.*?)</a>', re.S)
header_pat = re.compile(r'<a href="([^"]+)" class="url_header"', re.S)

def clean_href(href: str) -> str:
    href = unescape((href or '').strip())
    if href.startswith('//'):
        href = 'https:' + href
    if 'duckduckgo.com/l/?' in href:
        q = parse_qs(urlparse(href).query)
        if 'uddg' in q and q['uddg']:
            href = unquote(q['uddg'][0])
    return href


def is_bad(url: str) -> bool:
    host = urlparse(url).netloc.lower()
    return any(b in host for b in BAD)


def score(url: str, title: str, name: str, content: str) -> int:
    host = urlparse(url).netloc.lower()
    if not host or is_bad(url):
        return -999
    s = 0
    base = name.replace('（人工核对）', '').replace('（', '').replace('）', '')
    if base[:4] and base[:4] in title:
        s += 25
    if '官网' in title or '官方网站' in title or '首页' in title:
        s += 12
    if any(k in content for k in ['售前邮箱', '联系我们', '公司简介', '关于我们', '产品', '解决方案']):
        s += 5
    if host.count('.') <= 2:
        s += 3
    if host.endswith('.cn') or host.endswith('.com') or host.endswith('.vip'):
        s += 2
    return s


def pick_official(name: str) -> str:
    q = quote(f'{name} 官网')
    search_url = f'https://search.whyx.site:8444/search?q={q}'
    html = urlopen(Request(search_url, headers=UA), timeout=25).read().decode('utf-8', 'ignore')
    candidates = []
    for block in article_pat.findall(html):
        mm = url_pat.search(block)
        if not mm:
            continue
        href = clean_href(mm.group(1))
        title = re.sub(r'<.*?>', '', unescape(mm.group(2))).strip()
        content = re.sub(r'<.*?>', ' ', block)
        candidates.append((score(href, title, name, content), href, title))
    for href in header_pat.findall(html):
        href = clean_href(href)
        candidates.append((score(href, href, name, ''), href, href))
    candidates.sort(key=lambda x: (-x[0], len(urlparse(x[1]).netloc), x[1]))
    for sc, href, title in candidates:
        if sc > 0:
            return href
    return ''

# fill next 20 empty rows starting from row 2
filled_now = 0
for row in range(2, ws.max_row + 1):
    if filled_now >= 20:
        break
    if str(ws.cell(row, url_col).value or '').strip():
        continue
    name = str(ws.cell(row, 1).value or '').strip()
    if not name:
        continue
    try:
        url = pick_official(name)
    except Exception as e:
        url = ''
        print(f'ERR {row} {name} {e!r}', flush=True)
    if url:
        ws.cell(row, url_col).value = url
        wb.save(infile)
        print(f'{row}: {name} => {url}', flush=True)
        filled_now += 1
    else:
        print(f'{row}: {name} => [no confident url]', flush=True)
    time.sleep(0.25)

print('batch_filled', filled_now, flush=True)
